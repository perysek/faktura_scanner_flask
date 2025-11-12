"""
Serwis eksportu danych do Excel/CSV
"""
from pathlib import Path
from typing import List
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from database.models import Invoice


class ExportService:
	"""Serwis eksportu faktur"""
	
	def export_to_excel(self, invoices: List[Invoice], output_path: str):
		"""Eksportuj faktury do Excel"""
		wb = Workbook()
		ws = wb.active
		ws.title = "Faktury"
		
		# Header
		headers = [
			"ID", "Sprzedawca", "NIP", "Nr Faktury", "Data Faktury",
			"Konto Bankowe", "Kwota", "Waluta", "Termin Płatności", "Status",
			"Pewność OCR (%)", "Duplikat"
			]
		
		# Style dla header
		header_fill = PatternFill(
			start_color="4472C4", end_color="4472C4", fill_type="solid"
			)
		header_font = Font(color="FFFFFF", bold=True)
		
		for col, header in enumerate(headers, 1):
			cell = ws.cell(1, col, header)
			cell.fill = header_fill
			cell.font = header_font
			cell.alignment = Alignment(horizontal='center')
		
		# Dane
		for row, invoice in enumerate(invoices, 2):
			ws.cell(row, 1, invoice.id)
			ws.cell(row, 2, invoice.seller_name)
			ws.cell(row, 3, invoice.seller_nip or "")
			ws.cell(row, 4, invoice.invoice_number)
			ws.cell(
				row, 5, invoice.invoice_date.strftime(
					'%Y-%m-%d'
					) if invoice.invoice_date else ""
				)
			ws.cell(row, 6, invoice.bank_account or "")
			ws.cell(row, 7, invoice.amount)
			ws.cell(row, 8, invoice.currency)
			# Show payment_term if exists, otherwise payment_due_date
			payment_display = invoice.payment_term if invoice.payment_term else (
				invoice.payment_due_date.strftime('%Y-%m-%d') if invoice.payment_due_date else ""
			)
			ws.cell(row, 9, payment_display)
			ws.cell(row, 10, invoice.status)
			ws.cell(
				row, 11,
				f"{invoice.ocr_confidence:.1f}" if invoice.ocr_confidence else ""
				)
			ws.cell(row, 12, "TAK" if invoice.is_duplicate else "NIE")
		
		# Autowidth kolumn
		for col in ws.columns:
			max_length = 0
			col_letter = col[0].column_letter
			for cell in col:
				try:
					if len(str(cell.value)) > max_length:
						max_length = len(cell.value)
				except:
					pass
			adjusted_width = min(max_length + 2, 50)
			ws.column_dimensions[col_letter].width = adjusted_width
		
		# Zapisz
		wb.save(output_path)
	
	def export_to_csv(self, invoices: List[Invoice], output_path: str):
		"""Eksportuj faktury do CSV"""
		with open(output_path, 'w', newline='', encoding='utf-8-sig') as f:
			writer = csv.writer(f, delimiter=';')
			
			# Header
			writer.writerow(
				[
					"ID", "Sprzedawca", "NIP", "Nr Faktury", "Data Faktury",
					"Konto Bankowe", "Kwota", "Waluta", "Termin Płatności", "Status",
					"Pewność OCR (%)", "Duplikat"
					]
				)
			
			# Dane
			for invoice in invoices:
				# Show payment_term if exists, otherwise payment_due_date
				payment_display = invoice.payment_term if invoice.payment_term else (
					invoice.payment_due_date.strftime('%Y-%m-%d') if invoice.payment_due_date else ""
				)
				writer.writerow(
					[
						invoice.id,
						invoice.seller_name,
						invoice.seller_nip or "",
						invoice.invoice_number,
						invoice.invoice_date.strftime(
							'%Y-%m-%d'
							) if invoice.invoice_date else "",
						invoice.bank_account or "",
						invoice.amount,
						invoice.currency,
						payment_display,
						invoice.status,
						f"{invoice.ocr_confidence:.1f}" if invoice.ocr_confidence else "",
						"TAK" if invoice.is_duplicate else "NIE"
						]
					)