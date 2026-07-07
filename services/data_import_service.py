"""
Data import service — orchestrates the caldis.pl Playwright download +
xlsx parse + PostgreSQL INSERTs for the Flask app's import feature.

This is the in-app equivalent of scripts/import_appointments_playwright.py:
  - DB layer: PostgreSQL via psycopg2 (not SQLite)
  - Execution: synchronous from a background thread (not __main__)
  - Output: structured progress events via callback (not stdout)
  - Audit: writes to import_logs (not just printing summary)

Reuses fetch_xlsx_playwright() from the reference script unchanged.
"""
import asyncio
import json
import logging
from datetime import date, datetime
from pathlib import Path
from typing import Callable, List, Optional

import pandas as pd
import psycopg2.extensions

from config.database import get_pool
from exceptions import AppError
from repositories.data_import.import_log_repository import ImportLogRepository
from services.data_import_helpers import (
    DEFAULT_SERVICE_ID, KALENDARZ_OVERRIDES,
    build_employee_map, build_client_map, build_phone_map, build_service_map,
    resolve_employee_id, resolve_client_id, resolve_service_id,
    parse_client_name, create_client, normalize_phone,
    parse_appointment_date, parse_time, parse_created_at,
    calc_duration_minutes,
)

# Imported here so tests can patch it on this module's namespace
from scripts.import_appointments_playwright import fetch_xlsx_playwright

logger = logging.getLogger(__name__)


class ImportError(AppError):
    """Import pipeline failure — maps to HTTP 400."""
    status_code = 400


class DataImportService:
    """Orchestrates a single import run end-to-end."""

    def __init__(self, log_repo: Optional[ImportLogRepository] = None):
        self.log_repo = log_repo or ImportLogRepository()
        self.temp_dir = Path(__file__).resolve().parent.parent / 'assets' / 'temp'

    # ── entry point ──────────────────────────────────────────────────────────
    def run_import(self, import_id: int,
                   date_start: date, date_end: date,
                   dry_run: bool,
                   progress_callback: Callable[[dict], None],
                   keep_xlsx: bool = False) -> dict:
        """Run the full pipeline. Returns the final stats dict.

        Errors are caught, the log row is marked failed, and ImportError is raised
        so the runner can log it.
        """
        stats = self._zero_stats()
        pool = get_pool()
        xlsx_path: Optional[Path] = None
        # Three parallel lists — one per DB table — each row keyed by row_num
        row_data: dict = {'appointments': [], 'appointment_services': [], 'income_records': []}

        # Acquire DB connection upfront — before Playwright download — so
        # the failure handler always has a valid conn (no flask.g fallback
        # needed in a background thread).
        conn = pool.getconn()

        try:
            self._emit(progress_callback, 'log',
                       f"Start importu (zakres {date_start} → {date_end}, dry_run={dry_run})")

            # ── Phase 1: Playwright download ─────────────────────────────────
            xlsx_path = self._download_xlsx(import_id, date_start, date_end,
                                            progress_callback)

            # ── Phase 3: build lookup maps ───────────────────────────────────
            self._emit(progress_callback, 'log', 'Budowanie tablic wyszukiwania...')
            employee_map = build_employee_map(conn)
            client_map   = build_client_map(conn)
            phone_map    = build_phone_map(conn)
            service_list = build_service_map(conn)
            self._emit(progress_callback, 'log',
                       f"Pracownicy: {len(employee_map)}, "
                       f"klienci: {len(client_map) // 2}, "
                       f"klienci z telefonem: {len(phone_map)}, "
                       f"usługi: {len(service_list)}")

            # ── Phase 4: parse xlsx + insert ─────────────────────────────────
            df = pd.read_excel(xlsx_path, engine='openpyxl', dtype=str)
            df.columns = [c.strip() for c in df.columns]
            self._emit(progress_callback, 'log',
                       f"Wierszy do przetworzenia: {len(df)}")

            for idx, row in df.iterrows():
                self._process_row(row, idx, conn, dry_run, stats,
                                  employee_map, client_map, phone_map, service_list,
                                  progress_callback, row_data=row_data)
                processed = (stats['inserted'] + stats['skipped_zero']
                             + stats['skipped_no_client'] + stats['skipped_no_employee']
                             + stats['skipped_duplicate'] + stats['errors'])
                if processed % 10 == 0 and processed > 0:
                    self.log_repo.update_stats(import_id, stats, conn=conn)
                    self._emit(progress_callback, 'stats', None, stats=stats)

            # ── Phase 5: commit + mark completed ─────────────────────────────
            if not dry_run:
                conn.commit()
            else:
                conn.rollback()

            self.log_repo.mark_completed(import_id, stats, conn=conn)
            skipped = (stats['skipped_zero'] + stats['skipped_no_client']
                       + stats['skipped_no_employee'] + stats['skipped_duplicate'])

            # ── Phase 6: export XLSX audit file ─────────────────────────────
            xlsx_out = self._export_xlsx(
                row_data, import_id, date_start, date_end, dry_run, stats)
            self._emit(progress_callback, 'log',
                       f"XLSX zapisany: {xlsx_out.name}")

            self._emit(progress_callback, 'log',
                       f"Zakończono. Dodano: {stats['inserted']}, "
                       f"pominięto: {skipped}, błędy: {stats['errors']}")
            self._emit(progress_callback, 'status', None, status='completed')
            return stats

        except Exception as exc:
            logger.exception("Import %d failed", import_id)
            error_message = str(exc) or type(exc).__name__
            try:
                self.log_repo.mark_failed(import_id, error_message, stats=stats, conn=conn)
            except Exception:
                logger.exception("Could not write failure status to import_logs")
            if conn is not None:
                if 'Sesja wygasla' in error_message or 'Sesja wygas' in error_message:
                    try:
                        self.log_repo.update_session_status(import_id, 'expired', conn=conn)
                    except Exception:
                        pass
                elif 'Brak zapisanej sesji' in error_message:
                    try:
                        self.log_repo.update_session_status(import_id, 'missing', conn=conn)
                    except Exception:
                        pass
            self._emit(progress_callback, 'log', f"BŁĄD: {error_message}")
            self._emit(progress_callback, 'status', None, status='failed')
            raise ImportError(error_message) from exc

        finally:
            if conn is not None:
                try:
                    pool.putconn(conn)
                except Exception:
                    logger.exception("Could not return connection to pool")
            if xlsx_path is not None and xlsx_path.exists():
                if keep_xlsx:
                    try:
                        self._emit(progress_callback, 'log',
                                   f"Plik XLSX z caldis.pl zachowany: {xlsx_path}")
                    except Exception:
                        logger.warning("Could not emit keep-xlsx log for %s", xlsx_path)
                else:
                    try:
                        xlsx_path.unlink()
                    except Exception:
                        logger.warning("Could not delete xlsx %s", xlsx_path)

    # ── XLSX export ──────────────────────────────────────────────────────────
    def _export_xlsx(self, row_data: dict, import_id: int,
                     date_start: date, date_end: date,
                     dry_run: bool, stats: dict) -> Path:
        """Write a multi-sheet XLSX audit file to the project root.

        Sheets:
          1. Wizyty          — mirrors the `appointments` table (all rows + skipped)
          2. Uslugi_wizyt    — mirrors `appointment_services` (valid rows only)
          3. Przychody       — mirrors `income_records` (valid rows only)
          4. Podsumowanie    — import stats + metadata

        Cell types are Python native (int, float, datetime.date, datetime.datetime,
        bool) so Excel displays them correctly and PostgreSQL accepts them without
        additional casting. Skipped/error rows appear only in sheet 1 with
        action + skip_reason for auditability.

        File: caldis_import_<date_start>_<date_end>_<import_id>_<dryrun|real>.xlsx
        Path: project root (two levels above services/)
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        project_root = Path(__file__).resolve().parent.parent
        tag = 'dryrun' if dry_run else 'real'
        filename = (f"caldis_import_{date_start}_{date_end}"
                    f"_{import_id}_{tag}.xlsx")
        out_path = project_root / filename

        wb = Workbook()

        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(fill_type='solid', fgColor='1F4E79')

        def _make_sheet(title: str, headers: list, rows: List[dict]) -> None:
            ws = wb.create_sheet(title=title)
            ws.append(headers)
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
            ws.freeze_panes = 'A2'
            for r in rows:
                ws.append([r.get(h) for h in headers])
            # Auto-fit column widths (cap at 40)
            for col_idx, header in enumerate(headers, 1):
                max_len = max(
                    (len(str(ws.cell(row=r, column=col_idx).value or ''))
                     for r in range(1, ws.max_row + 1)),
                    default=len(header),
                )
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 40)

        # ── Sheet 1: Wizyty (appointments) ────────────────────────────────
        wb.remove(wb.active)  # remove default sheet
        _make_sheet('Wizyty', [
            'row_num', 'action', 'skip_reason',
            'client_id', 'employee_id', 'status',
            'appointment_date', 'start_time', 'end_time',
            'total_price', 'total_duration', 'discount_amount',
            'created_at', 'updated_at', 'is_deleted',
            'raw_name', 'raw_kalendarz', 'raw_kategoria', 'dry_run',
        ], row_data.get('appointments', []))

        # ── Sheet 2: Uslugi_wizyt (appointment_services) ──────────────────
        _make_sheet('Uslugi_wizyt', [
            'row_num',
            'appointment_id', 'service_id',
            'price_charged', 'duration_minutes',
            'commission_rate', 'commission_amount', 'is_addon',
        ], row_data.get('appointment_services', []))

        # ── Sheet 3: Przychody (income_records) ───────────────────────────
        _make_sheet('Przychody', [
            'row_num',
            'appointment_id', 'client_id', 'employee_id',
            'total_amount', 'discount_amount', 'net_amount', 'commission_total',
            'payment_date', 'created_at',
        ], row_data.get('income_records', []))

        # ── Sheet 4: Podsumowanie (summary) ──────────────────────────────
        ws_sum = wb.create_sheet(title='Podsumowanie')
        summary_rows = [
            ('import_id', import_id),
            ('date_range_start', date_start),
            ('date_range_end', date_end),
            ('dry_run', dry_run),
            ('generated_at', datetime.now()),
            ('', ''),
            ('inserted', stats.get('inserted', 0)),
            ('clients_created', stats.get('clients_created', 0)),
            ('skipped_zero', stats.get('skipped_zero', 0)),
            ('skipped_no_client', stats.get('skipped_no_client', 0)),
            ('skipped_no_employee', stats.get('skipped_no_employee', 0)),
            ('skipped_duplicate', stats.get('skipped_duplicate', 0)),
            ('errors', stats.get('errors', 0)),
            # clients_created is a subset of inserted rows, not a distinct row
            # category — exclude it so total_rows still equals rows processed.
            ('total_rows', sum(v for k, v in stats.items() if k != 'clients_created')),
        ]
        for label, value in summary_rows:
            row = ws_sum.append([label, value])
        ws_sum.column_dimensions['A'].width = 22
        ws_sum.column_dimensions['B'].width = 28
        for cell in ws_sum['A']:
            if cell.value:
                cell.font = Font(bold=True)

        wb.save(out_path)
        logger.info("XLSX audit written: %s (%d appt rows, %d svc rows, %d income rows)",
                    out_path, len(row_data.get('appointments', [])),
                    len(row_data.get('appointment_services', [])),
                    len(row_data.get('income_records', [])))
        return out_path

    # ── helpers ──────────────────────────────────────────────────────────────
    @staticmethod
    def _zero_stats() -> dict:
        return {
            'inserted': 0,
            'clients_created': 0,
            'skipped_zero': 0,
            'skipped_no_client': 0,
            'skipped_no_employee': 0,
            'skipped_duplicate': 0,
            'errors': 0,
        }

    @staticmethod
    def _emit(callback: Callable[[dict], None], event_type: str,
              message: Optional[str], **extra) -> None:
        """Build a progress event dict and call the callback."""
        event = {
            'type': event_type,
            'message': message,
            'timestamp': datetime.now().isoformat(),
            **extra,
        }
        callback(event)

    def _download_xlsx(self, import_id: int, date_start: date, date_end: date,
                       progress_callback: Callable[[dict], None]) -> Path:
        """Run Playwright download, update session_status accordingly."""
        self.temp_dir.mkdir(parents=True, exist_ok=True)
        xlsx_name = f"caldis_pw_rezerwacje_{date_start}_{date_end}_{import_id}.xlsx"
        xlsx_path = self.temp_dir / xlsx_name
        self._emit(progress_callback, 'log', 'Pobieranie xlsx z caldis.pl (Playwright)...')
        asyncio.run(
            fetch_xlsx_playwright(
                email=None, password=None,
                date_start=date_start, date_end=date_end,
                output_path=xlsx_path,
                headed=False,
            )
        )
        try:
            self.log_repo.update_session_status(import_id, 'active')
        except Exception:
            pass
        self._emit(progress_callback, 'log', f"Pobrano: {xlsx_path.name}")
        return xlsx_path

    def _process_row(self, row, idx: int,
                     conn: psycopg2.extensions.connection,
                     dry_run: bool, stats: dict,
                     employee_map: dict, client_map: dict,
                     phone_map: dict, service_list: list,
                     progress_callback: Callable[[dict], None],
                     row_data: Optional[dict] = None) -> None:
        """Parse + dedupe + insert a single xlsx row. Updates stats in place.

        Appends typed dicts to row_data['appointments'], ['appointment_services'],
        ['income_records'] for every row so the XLSX export captures all FK
        relationships and cross-reference tables.

        All values are stored as Python native types (int, float, datetime.date,
        datetime.datetime, bool) — not strings — so openpyxl writes them with
        correct Excel cell types, and PostgreSQL accepts them without casting.
        """
        raw_name      = str(row.get('Imię i nazwisko', row.get('Imie i nazwisko', '')))
        raw_kalendarz = str(row.get('Kalendarz', ''))
        raw_kategoria = str(row.get('Kategoria', ''))

        def _appt(action: str, skip_reason: str = '', **extra) -> None:
            """Append to the appointments audit sheet."""
            if row_data is None:
                return
            # Parse appointment_date to datetime.date for typed Excel cell
            appt_dt = None
            if extra.get('appointment_date'):
                try:
                    from datetime import date as _date
                    appt_dt = _date.fromisoformat(extra['appointment_date'])
                except Exception:
                    appt_dt = extra['appointment_date']
            # Parse created_at to datetime for typed Excel cell
            created_dt = None
            if extra.get('created_at'):
                try:
                    created_dt = datetime.strptime(extra['created_at'], '%Y-%m-%d %H:%M:%S')
                except Exception:
                    created_dt = extra.get('created_at')
            row_data['appointments'].append({
                'row_num':          int(idx),
                'action':           action,
                'skip_reason':      skip_reason,
                # ── appointments table columns ───────────────────────────────
                'client_id':        extra.get('client_id'),     # INTEGER FK → clients.id
                'employee_id':      extra.get('employee_id'),   # INTEGER FK → employees.id
                # Mirror the real INSERT: future visits → 'scheduled', else 'completed'.
                'status':           (('scheduled' if (extra.get('appointment_date') or '') > date.today().isoformat()
                                       else 'completed') if action == 'inserted' else None),
                'appointment_date': appt_dt,                    # DATE → YYYY-MM-DD
                'start_time':       extra.get('start_time'),    # TIME → HH:MM:SS string
                'end_time':         extra.get('end_time'),      # TIME → HH:MM:SS string
                'total_price':      extra.get('total_price'),   # NUMERIC
                'total_duration':   extra.get('duration_minutes'), # INTEGER (minutes)
                'discount_amount':  0.0,                        # NUMERIC NOT NULL DEFAULT 0
                'created_at':       created_dt,                 # TIMESTAMP
                'updated_at':       created_dt,                 # TIMESTAMP (same as created)
                'is_deleted':       False,                      # BOOLEAN NOT NULL DEFAULT FALSE
                # ── source columns (audit metadata) ─────────────────────────
                'raw_name':         raw_name,
                'raw_kalendarz':    raw_kalendarz,
                'raw_kategoria':    raw_kategoria,
                'dry_run':          dry_run,
            })

        def _appt_svc(appointment_id, service_id, total_price,
                      duration_minutes, commission_rate, commission_amount) -> None:
            """Append to the appointment_services audit sheet."""
            if row_data is None:
                return
            row_data['appointment_services'].append({
                'row_num':           int(idx),
                # ── appointment_services table columns ───────────────────────
                'appointment_id':    appointment_id,   # INTEGER FK → appointments.id (NULL=dry_run)
                'service_id':        int(service_id),  # INTEGER FK → services.id
                'price_charged':     float(total_price),     # NUMERIC
                'duration_minutes':  int(duration_minutes),  # INTEGER
                'commission_rate':   float(commission_rate), # NUMERIC
                'commission_amount': float(commission_amount), # NUMERIC
                'is_addon':          False,             # BOOLEAN NOT NULL DEFAULT FALSE
            })

        def _income(appointment_id, client_id, employee_id,
                    total_price, commission_amount, appointment_date, created_at) -> None:
            """Append to the income_records audit sheet."""
            if row_data is None:
                return
            appt_dt = None
            if appointment_date:
                try:
                    from datetime import date as _date
                    appt_dt = _date.fromisoformat(appointment_date)
                except Exception:
                    appt_dt = appointment_date
            created_dt = None
            if created_at:
                try:
                    created_dt = datetime.strptime(created_at, '%Y-%m-%d %H:%M:%S')
                except Exception:
                    created_dt = created_at
            row_data['income_records'].append({
                'row_num':          int(idx),
                # ── income_records table columns ─────────────────────────────
                'appointment_id':   appointment_id,       # INTEGER FK → appointments.id
                # client_id is None only for a dry-run row that would create a
                # brand-new client — nothing was actually inserted to cast.
                'client_id':        (int(client_id) if client_id is not None else None),
                'employee_id':      int(employee_id),     # INTEGER FK → employees.id
                'total_amount':     float(total_price),   # NUMERIC
                'discount_amount':  0.0,                  # NUMERIC NOT NULL DEFAULT 0
                'net_amount':       float(total_price),   # NUMERIC (gross = net here)
                'commission_total': float(commission_amount), # NUMERIC
                'payment_date':     appt_dt,              # DATE
                'created_at':       created_dt,           # TIMESTAMP
            })

        try:
            suma_brutto_raw = row.get('Suma brutto', '0')
            try:
                suma_brutto = float(str(suma_brutto_raw).replace(',', '.')) if suma_brutto_raw else 0.0
            except (ValueError, TypeError):
                suma_brutto = 0.0
            if suma_brutto == 0:
                stats['skipped_zero'] += 1
                _appt('skipped_zero', 'Suma brutto = 0')
                return

            od_cell       = row.get('Od', '')
            do_cell       = row.get('Do', '')
            imie_cell     = row.get('Imię i nazwisko', row.get('Imie i nazwisko', ''))
            telefon_cell  = row.get('Telefon', '')
            kalendarz_cell = row.get('Kalendarz', '')
            kategoria_cell = row.get('Kategoria', '')

            appointment_date = parse_appointment_date(od_cell)
            start_time       = parse_time(od_cell)
            end_time         = parse_time(do_cell)
            created_at       = parse_created_at(row.get('Data utworzenia', ''))
            duration_minutes = calc_duration_minutes(od_cell, do_cell)

            if not (appointment_date and start_time and end_time):
                stats['errors'] += 1
                _appt('error', f'Brak daty/czasu: od={od_cell!r} do={do_cell!r}')
                return

            kal_lower = str(kalendarz_cell).strip().lower() if kalendarz_cell else ''
            if kal_lower in KALENDARZ_OVERRIDES:
                employee_id, forced_service_id = KALENDARZ_OVERRIDES[kal_lower]
            else:
                employee_id = resolve_employee_id(kalendarz_cell, employee_map)
                forced_service_id = None

            if employee_id is None:
                stats['skipped_no_employee'] += 1
                _appt('skipped_no_employee',
                      f'Nie znaleziono pracownika: {kalendarz_cell!r}',
                      appointment_date=appointment_date, start_time=start_time)
                return

            client_id = resolve_client_id(imie_cell, client_map, telefon_cell, phone_map)
            new_client_note = ''
            pending_new_client = None
            if client_id is None:
                parsed_name = parse_client_name(imie_cell)
                if parsed_name is None:
                    # Blank cell or the 'Wolne' placeholder (blocked calendar
                    # slot) — not a client, nothing to import.
                    stats['skipped_no_client'] += 1
                    _appt('skipped_no_client',
                          f'Brak nazwiska klienta (puste pole lub "Wolne"): {imie_cell!r}',
                          appointment_date=appointment_date, start_time=start_time,
                          employee_id=employee_id)
                    return
                # A real name that matches no existing client. Hold off on
                # actually creating it until after the duplicate check below,
                # so a row that turns out to be a duplicate slot never leaves
                # an orphan client behind.
                pending_new_client = parsed_name

            # Duplicate check
            cursor = conn.cursor()
            cursor.execute(
                """
                SELECT 1 FROM appointments
                WHERE employee_id = %s AND appointment_date = %s AND start_time = %s
                  AND is_deleted = FALSE
                """,
                (employee_id, appointment_date, start_time),
            )
            if cursor.fetchone() is not None:
                stats['skipped_duplicate'] += 1
                _appt('skipped_duplicate',
                      f'Duplikat: employee_id={employee_id} '
                      f'date={appointment_date} time={start_time}',
                      appointment_date=appointment_date, start_time=start_time,
                      end_time=end_time, client_id=client_id,
                      employee_id=employee_id, duration_minutes=duration_minutes,
                      created_at=created_at)
                return

            if pending_new_client is not None:
                # caldis.pl is the source of truth for new customers here — a
                # name that matches no existing client is a new client, not a
                # data error, so create one instead of skipping the visit.
                first_name, last_name = pending_new_client
                client_phone = normalize_phone(telefon_cell)
                stats['clients_created'] += 1
                new_client_note = f'Nowy klient: {first_name} {last_name}'.strip()

                if dry_run:
                    client_id = None  # nothing written yet — simulated only
                else:
                    client_id = create_client(conn, first_name, last_name, client_phone)
                    # Repeat rows for this same new client later in the same
                    # file now resolve to the row just created, not a duplicate.
                    client_map[(first_name.lower(), last_name.lower())] = client_id
                    if client_phone:
                        phone_map[client_phone] = client_id
                    self._emit(progress_callback, 'log',
                               f'Nowy klient utworzony: {first_name} {last_name} (id={client_id})')

            service_id = (forced_service_id if forced_service_id is not None
                          else resolve_service_id(kategoria_cell, service_list))

            cursor.execute("SELECT commission_rate FROM employees WHERE id = %s",
                           (employee_id,))
            emp_row = cursor.fetchone()
            commission_rate = float(emp_row['commission_rate'] or 0) if emp_row else 0.0
            commission_amount = round(suma_brutto * commission_rate / 100, 2)
            total_price = round(suma_brutto, 2)

            # Future-dated imports are upcoming visits, not historical ones:
            # mark them 'scheduled' so they appear as active bookings in the
            # calendar. Past/today appointments keep 'completed'. appointment_date
            # is an ISO 'YYYY-MM-DD' string, so lexical comparison == chronological.
            appt_status = 'scheduled' if appointment_date > date.today().isoformat() else 'completed'
            is_future = appt_status == 'scheduled'

            if dry_run:
                stats['inserted'] += 1
                _appt('inserted', new_client_note,
                      appointment_date=appointment_date, start_time=start_time,
                      end_time=end_time, duration_minutes=duration_minutes,
                      created_at=created_at,
                      client_id=(int(client_id) if client_id is not None else None),
                      employee_id=int(employee_id), total_price=total_price)
                _appt_svc(None, service_id, total_price,
                          duration_minutes, commission_rate, commission_amount)
                # Future visits haven't generated revenue yet — no income record.
                if not is_future:
                    _income(None, client_id, employee_id,
                            total_price, commission_amount, appointment_date, created_at)
                return

            # INSERT appointments
            cursor.execute(
                """
                INSERT INTO appointments (
                    client_id, employee_id, status,
                    appointment_date, start_time, end_time,
                    total_price, total_duration, discount_amount,
                    created_at, updated_at
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, 0, %s, %s)
                RETURNING id
                """,
                (client_id, employee_id, appt_status,
                 appointment_date, start_time, end_time,
                 total_price, duration_minutes,
                 created_at, created_at),
            )
            appointment_id = cursor.fetchone()['id']

            # INSERT appointment_services
            cursor.execute(
                """
                INSERT INTO appointment_services (
                    appointment_id, service_id, price_charged, duration_minutes,
                    commission_rate, commission_amount, is_addon
                ) VALUES (%s, %s, %s, %s, %s, %s, FALSE)
                """,
                (appointment_id, service_id, total_price, duration_minutes,
                 commission_rate, commission_amount),
            )

            # Income + last-visit are past-tense facts: only record them for
            # completed (past/today) visits. A future 'scheduled' appointment
            # hasn't happened yet, so recognizing revenue or stamping
            # last_visit_date in the future would be wrong.
            if not is_future:
                # INSERT income_records
                cursor.execute(
                    """
                    INSERT INTO income_records (
                        appointment_id, client_id, employee_id,
                        total_amount, discount_amount, net_amount, commission_total,
                        payment_date, created_at
                    ) VALUES (%s, %s, %s, %s, 0, %s, %s, %s, %s)
                    """,
                    (appointment_id, client_id, employee_id,
                     total_price, total_price, commission_amount,
                     appointment_date, created_at),
                )

                # UPDATE clients.last_visit_date
                cursor.execute(
                    """
                    UPDATE clients
                    SET last_visit_date = %s
                    WHERE id = %s
                      AND (last_visit_date IS NULL OR last_visit_date < %s)
                    """,
                    (appointment_date, client_id, appointment_date),
                )

            stats['inserted'] += 1
            _appt('inserted', new_client_note,
                  appointment_date=appointment_date, start_time=start_time,
                  end_time=end_time, duration_minutes=duration_minutes,
                  created_at=created_at, client_id=int(client_id),
                  employee_id=int(employee_id), total_price=total_price)
            _appt_svc(int(appointment_id), service_id, total_price,
                      duration_minutes, commission_rate, commission_amount)
            if not is_future:
                _income(int(appointment_id), client_id, employee_id,
                        total_price, commission_amount, appointment_date, created_at)

        except Exception:
            logger.exception("Error processing row %d", idx)
            stats['errors'] += 1
            _appt('error', f'Exception na wierszu {idx}')
